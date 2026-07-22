"""Spreadsheet / structured-file readers for the import lanes.

Reads xlsx, csv, and yaml into lists of row-dicts, then maps them onto the
account and trekker schemas. Column names are matched loosely (case/space
insensitive) so real-world sheets import without hand-editing headers.
"""
from __future__ import annotations

import csv
import io
from typing import Dict, List, Optional

import yaml
from openpyxl import load_workbook

from app.import_.parser import _norm_gender, _norm_mobile  # reuse normalizers
from app.portal.ids import detect_id_type, to_canonical


def _norm_key(k: str) -> str:
    return "".join(ch for ch in str(k).strip().lower() if ch.isalnum())


# canonical field -> accepted header spellings (normalized)
_ACCOUNT_HEADERS = {
    "email": {"email", "emailid", "mail", "username", "user"},
    "password": {"password", "pass", "pwd"},
    "status": {"status", "state"},
    "notes": {"notes", "note", "remark", "remarks"},
}
_TREKKER_HEADERS = {
    "name": {"name", "fullname", "trekker", "passenger", "person"},
    "age": {"age", "years", "yrs"},
    "gender": {"gender", "sex"},
    "mobile_no": {"mobileno", "mobile", "phone", "phonenumber", "contact",
                  "contactnumber", "mob", "whatsapp"},
    "govt_id_type": {"govtidtype", "idtype", "documenttype"},
    "govt_id": {"govtid", "id", "idno", "idnumber", "pan", "aadhaar", "voterid",
                "dl", "govtidno", "idproof"},
}


def _rows_from_bytes(data: bytes, filename: str) -> List[Dict]:
    """Return a list of row-dicts keyed by the file's headers."""
    name = (filename or "").lower()
    if name.endswith((".yaml", ".yml")):
        return _rows_from_yaml(data)
    if name.endswith(".csv") or name.endswith(".tsv"):
        return _rows_from_csv(data, tsv=name.endswith(".tsv"))
    if name.endswith((".xlsx", ".xlsm")):
        return _rows_from_xlsx(data)
    # Fallback: try csv, then yaml.
    try:
        rows = _rows_from_csv(data)
        if rows:
            return rows
    except Exception:
        pass
    return _rows_from_yaml(data)


def _rows_from_csv(data: bytes, tsv: bool = False) -> List[Dict]:
    text = data.decode("utf-8-sig", errors="replace")
    delim = "\t" if tsv else None
    if delim is None:
        # sniff comma vs semicolon vs tab
        sample = text[:2000]
        delim = ";" if sample.count(";") > sample.count(",") else ","
        if sample.count("\t") > sample.count(delim):
            delim = "\t"
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    return [dict(r) for r in reader]


def _rows_from_xlsx(data: bytes) -> List[Dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


def _rows_from_yaml(data: bytes) -> List[Dict]:
    doc = yaml.safe_load(data.decode("utf-8", errors="replace"))
    if isinstance(doc, dict):
        # accept {accounts: [...]}, {trekkers: [...]} or a single mapping
        for key in ("accounts", "trekkers", "rows", "items"):
            if isinstance(doc.get(key), list):
                return [d for d in doc[key] if isinstance(d, dict)]
        return [doc]
    if isinstance(doc, list):
        return [d for d in doc if isinstance(d, dict)]
    return []


def _map_row(row: Dict, header_map: Dict[str, set]) -> Dict[str, Optional[str]]:
    """Map a raw row-dict onto canonical fields using the header spellings."""
    normalized = {_norm_key(k): v for k, v in row.items()}
    out: Dict[str, Optional[str]] = {}
    for field, spellings in header_map.items():
        val = None
        for nk, v in normalized.items():
            if nk in spellings:
                val = v
                break
        out[field] = val
    return out


def read_accounts(data: bytes, filename: str) -> List[Dict]:
    rows = _rows_from_bytes(data, filename)
    accounts = []
    for r in rows:
        m = _map_row(r, _ACCOUNT_HEADERS)
        email = (str(m["email"]).strip().lower() if m["email"] else "")
        if not email or "@" not in email:
            continue
        status = str(m["status"]).strip().lower() if m["status"] else "available"
        if status not in ("available", "booked", "disabled"):
            status = "available"
        accounts.append({
            "email": email,
            "password": str(m["password"]).strip() if m["password"] else None,
            "status": status,
            "notes": str(m["notes"]).strip() if m["notes"] else None,
        })
    return accounts


def read_trekkers(data: bytes, filename: str) -> List[Dict]:
    rows = _rows_from_bytes(data, filename)
    trekkers = []
    for r in rows:
        m = _map_row(r, _TREKKER_HEADERS)
        name = str(m["name"]).strip() if m["name"] else None
        if not name:
            continue
        gid = str(m["govt_id"]).strip() if m["govt_id"] else None
        gtype = to_canonical(m["govt_id_type"]) if m["govt_id_type"] else None
        if gid and not gtype:
            gtype = detect_id_type(gid)
        age = None
        if m["age"] not in (None, ""):
            try:
                age = int(float(str(m["age"]).strip()))
            except ValueError:
                age = None
        rec = {
            "name": name,
            "age": age,
            "gender": _norm_gender(str(m["gender"])) if m["gender"] else None,
            "mobile_no": _norm_mobile(str(m["mobile_no"])) if m["mobile_no"] else None,
            "govt_id_type": gtype,
            "govt_id": gid.upper() if gid else None,
        }
        rec["issues"] = [f for f in ("name", "age", "gender", "mobile_no",
                                      "govt_id_type", "govt_id") if not rec[f]]
        trekkers.append(rec)
    return trekkers
